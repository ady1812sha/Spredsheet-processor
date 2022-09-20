import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def spreadsheet_processing(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    new_cell = sheet.cell(1, sheet.max_column)
    new_cell.value = "Corrected price"
    chart_values = Reference(sheet,
                             min_row=2,
                             max_row=sheet.max_row,
                             min_col=4,
                             max_col=4)
    chart = BarChart()
    chart.add_data(chart_values)
    sheet.add_chart(chart, "e2")
    wb.save(filename)

filename = input("which file do you want to process? ")
spreadsheet_processing(filename)